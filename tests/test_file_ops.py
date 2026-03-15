import json
import subprocess
import tempfile
import unittest
import zipfile
from pathlib import Path

from PIL import Image
from openpyxl import Workbook
import fitz


ROOT = Path(__file__).resolve().parents[1]
SKILL_DIR = ROOT / "skills" / "file-ops"
SCRIPT = SKILL_DIR / "scripts" / "file_ops.py"

# Prefer skill-local .venv; fall back to repo root .venv for development
_skill_python = SKILL_DIR / ".venv" / "bin" / "python"
PYTHON = _skill_python if _skill_python.exists() else ROOT / ".venv" / "bin" / "python"


def run_command(*args: str) -> tuple[int, dict]:
    completed = subprocess.run(
        [str(PYTHON), str(SCRIPT), *args],
        capture_output=True,
        text=True,
        check=False,
    )
    return completed.returncode, json.loads(completed.stdout)


class HealthTests(unittest.TestCase):
    def test_health_reports_expected_keys(self) -> None:
        code, payload = run_command("health")
        self.assertEqual(code, 0)
        self.assertTrue(payload["success"])
        self.assertIn("available_conversions", payload)
        self.assertIn("available_operations", payload)
        self.assertIn("runtime", payload)
        self.assertIn("image -> image", payload["available_conversions"])

    def test_health_reports_new_runtime_fields(self) -> None:
        code, payload = run_command("health")
        self.assertEqual(code, 0)
        self.assertIn("python_docx", payload["runtime"])
        self.assertIn("pymupdf", payload["runtime"])


class ConvertTests(unittest.TestCase):
    def test_image_conversion_png_to_jpg(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.png"
            Image.new("RGBA", (12, 12), (255, 0, 0, 128)).save(source)

            code, payload = run_command("convert", "--input", str(source), "--to", "jpg")
            self.assertEqual(code, 0, payload)

            output = Path(payload["output_path"])
            self.assertTrue(output.exists())
            self.assertEqual(output.suffix.lower(), ".jpg")

    def test_pdf_to_docx_conversion(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.pdf"
            doc = fitz.open()
            page = doc.new_page()
            page.insert_text((72, 72), "Hello from file-ops")
            doc.save(source)
            doc.close()

            code, payload = run_command("convert", "--input", str(source), "--to", "docx")
            self.assertEqual(code, 0, payload)

            output = Path(payload["output_path"])
            self.assertTrue(output.exists())
            self.assertEqual(output.suffix.lower(), ".docx")

    def test_excel_to_csv_conversion(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["name", "value"])
            ws.append(["alpha", 1])
            wb.save(source)

            code, payload = run_command("convert", "--input", str(source), "--to", "csv")
            self.assertEqual(code, 0, payload)

            output = Path(payload["output_path"])
            self.assertTrue(output.exists())
            self.assertIn("alpha,1", output.read_text(encoding="utf-8"))

    def test_docx_to_pdf_fails_cleanly_without_libreoffice(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.docx"
            source.write_bytes(b"not-a-real-docx")

            code, payload = run_command("convert", "--input", str(source), "--to", "pdf")
            self.assertEqual(code, 1)
            self.assertIn("LibreOffice", payload["error"])


class InspectTests(unittest.TestCase):
    def test_inspect_image(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.png"
            Image.new("RGB", (64, 32)).save(source)

            code, payload = run_command("inspect", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertTrue(payload["success"])
            self.assertIn("size_bytes", payload)
            self.assertIn("mime_type", payload)
            self.assertEqual(payload["dimensions"]["width"], 64)
            self.assertEqual(payload["dimensions"]["height"], 32)

    def test_inspect_pdf(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.pdf"
            doc = fitz.open()
            doc.new_page()
            doc.new_page()
            doc.save(source)
            doc.close()

            code, payload = run_command("inspect", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertEqual(payload["page_count"], 2)

    def test_inspect_excel(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["a", "b"])
            ws.append(["c", "d"])
            wb.save(source)

            code, payload = run_command("inspect", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertEqual(payload["sheet_names"], ["Sheet1"])
            self.assertEqual(payload["row_counts"]["Sheet1"], 2)


class ArchiveTests(unittest.TestCase):
    def test_archive_create_and_extract(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)

            # Create source files
            src = tmp_path / "src"
            src.mkdir()
            (src / "a.txt").write_text("hello")
            (src / "b.txt").write_text("world")

            archive_path = tmp_path / "out.zip"

            # Create archive
            code, payload = run_command(
                "archive", "--input", str(src), "--output", str(archive_path)
            )
            self.assertEqual(code, 0, payload)
            self.assertTrue(archive_path.exists())
            self.assertEqual(payload["files_added"], 2)

            # Extract archive
            dest = tmp_path / "extracted"
            code, payload = run_command(
                "archive", "--input", str(archive_path), "--output", str(dest), "--extract"
            )
            self.assertEqual(code, 0, payload)
            self.assertTrue(dest.exists())
            self.assertGreater(payload["entries"], 0)

    def test_archive_rejects_non_zip_output(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "a.txt"
            source.write_text("hello")
            code, payload = run_command(
                "archive", "--input", str(source), "--output", str(Path(tmp) / "out.tar")
            )
            self.assertEqual(code, 1)
            self.assertIn(".zip", payload["error"])

    def test_archive_zipslip_protection(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)

            # Create a malicious zip with path traversal entry
            malicious_zip = tmp_path / "evil.zip"
            with zipfile.ZipFile(malicious_zip, "w") as zf:
                zf.writestr("../../../etc/evil.txt", "pwned")

            dest = tmp_path / "extracted"
            code, payload = run_command(
                "archive", "--input", str(malicious_zip), "--output", str(dest), "--extract"
            )
            self.assertEqual(code, 1)
            self.assertIn("path traversal", payload["error"])

    def test_archive_handles_duplicate_basenames(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            dir_a = tmp_path / "a"
            dir_b = tmp_path / "b"
            dir_a.mkdir()
            dir_b.mkdir()
            (dir_a / "report.txt").write_text("from a")
            (dir_b / "report.txt").write_text("from b")

            archive_path = tmp_path / "out.zip"
            code, payload = run_command(
                "archive",
                "--input", str(dir_a / "report.txt"), str(dir_b / "report.txt"),
                "--output", str(archive_path),
            )
            self.assertEqual(code, 0, payload)

            # Both files should be present (no silent overwrite)
            with zipfile.ZipFile(archive_path) as zf:
                self.assertEqual(len(zf.namelist()), 2)


class ExtractTextTests(unittest.TestCase):
    def test_extract_text_from_pdf(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.pdf"
            doc = fitz.open()
            page = doc.new_page()
            page.insert_text((72, 72), "Extract this text")
            doc.save(source)
            doc.close()

            code, payload = run_command("extract-text", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertIn("Extract this text", payload["text"])
            self.assertEqual(payload["format"], "pdf")

    def test_extract_text_from_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["key", "val"])
            ws.append(["foo", "bar"])
            wb.save(source)

            code, payload = run_command("extract-text", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertIn("foo", payload["text"])
            self.assertIn("bar", payload["text"])

    def test_extract_text_from_html(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.html"
            source.write_text(
                "<html><body><p>Hello HTML</p><script>ignore</script></body></html>",
                encoding="utf-8",
            )

            code, payload = run_command("extract-text", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertIn("Hello HTML", payload["text"])
            self.assertNotIn("ignore", payload["text"])

    def test_extract_text_from_markdown(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.md"
            source.write_text("# Title\n\nParagraph text.", encoding="utf-8")

            code, payload = run_command("extract-text", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertIn("Paragraph text", payload["text"])

    def test_extract_text_from_docx(self) -> None:
        from docx import Document

        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.docx"
            doc = Document()
            doc.add_paragraph("Hello from python-docx")
            doc.save(str(source))

            code, payload = run_command("extract-text", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertIn("Hello from python-docx", payload["text"])
            self.assertEqual(payload["format"], "docx")

    def test_extract_text_from_image_no_exif(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.png"
            Image.new("RGB", (10, 10)).save(source)

            code, payload = run_command("extract-text", "--input", str(source))
            self.assertEqual(code, 0)
            self.assertIn("No EXIF", payload["text"])

    def test_extract_text_unsupported_format(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "sample.xyz"
            source.write_text("data")

            code, payload = run_command("extract-text", "--input", str(source))
            self.assertEqual(code, 1)
            self.assertIn("Unsupported", payload["error"])


if __name__ == "__main__":
    unittest.main()
