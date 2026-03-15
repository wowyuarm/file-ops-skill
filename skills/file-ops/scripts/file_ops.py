#!/usr/bin/env python3
"""Local file operations skill entry point.

Provides commands: health, convert, inspect, archive, extract-text.
All commands return structured JSON to stdout.
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import mimetypes
import shutil
import stat
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

IMAGE_FORMATS = {"jpg", "jpeg", "png", "webp", "gif", "bmp", "tiff"}
EXCEL_FORMATS = {"xlsx", "xlsm", "xltx", "xltm"}
LEGACY_EXCEL_FORMATS = {"xls"}
HTML_FORMATS = {"html", "htm"}
MARKDOWN_FORMATS = {"md", "markdown"}


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class FileOpsError(Exception):
    """User-facing error with an actionable message."""


# ---------------------------------------------------------------------------
# HTML text extractor (stdlib only)
# ---------------------------------------------------------------------------


class _HTMLTextExtractor(HTMLParser):
    """Strip tags from HTML and return readable text."""

    _IGNORED_TAGS = {"script", "style"}
    _BLOCK_TAGS = {"br", "div", "p", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._chunks: list[str] = []
        self._ignore_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in self._IGNORED_TAGS:
            self._ignore_depth += 1
        elif tag in self._BLOCK_TAGS:
            self._chunks.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in self._IGNORED_TAGS and self._ignore_depth > 0:
            self._ignore_depth -= 1
        elif tag in self._BLOCK_TAGS:
            self._chunks.append("\n")

    def handle_data(self, data: str) -> None:
        if self._ignore_depth:
            return
        text = data.strip()
        if text:
            self._chunks.append(text)

    def get_text(self) -> str:
        raw = "".join(self._chunks)
        return "\n".join(line.strip() for line in raw.splitlines() if line.strip())


# ---------------------------------------------------------------------------
# Runtime detection
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class RuntimeStatus:
    libreoffice: str | None
    wkhtmltopdf: str | None
    openpyxl: bool
    xlrd: bool
    markdown: bool
    pdfkit: bool
    python_docx: bool
    pymupdf: bool


def _has_module(name: str) -> bool:
    return importlib.util.find_spec(name) is not None


def detect_runtime() -> RuntimeStatus:
    return RuntimeStatus(
        libreoffice=shutil.which("soffice") or shutil.which("libreoffice"),
        wkhtmltopdf=shutil.which("wkhtmltopdf"),
        openpyxl=_has_module("openpyxl"),
        xlrd=_has_module("xlrd"),
        markdown=_has_module("markdown"),
        pdfkit=_has_module("pdfkit"),
        python_docx=_has_module("docx"),
        pymupdf=_has_module("fitz"),
    )


def build_health_report() -> dict:
    runtime = detect_runtime()

    available_conversions: list[str] = ["image -> image", "pdf -> docx"]
    unavailable_conversions: list[str] = []

    if runtime.openpyxl or runtime.xlrd:
        available_conversions.append("excel -> csv")
    else:
        unavailable_conversions.append("excel -> csv (missing openpyxl/xlrd)")

    if runtime.libreoffice:
        available_conversions.append("docx -> pdf")
    else:
        unavailable_conversions.append("docx -> pdf (missing LibreOffice)")

    if runtime.wkhtmltopdf and runtime.markdown and runtime.pdfkit:
        available_conversions.extend(["html -> pdf", "markdown -> pdf"])
    else:
        unavailable_conversions.append(
            "html/markdown -> pdf (missing wkhtmltopdf or python helpers)"
        )

    available_ops: list[str] = [
        "inspect -> basic file metadata",
        "archive -> create/extract zip",
        "extract-text -> html",
        "extract-text -> markdown",
    ]
    unavailable_ops: list[str] = []

    if runtime.openpyxl:
        available_ops.extend(["inspect -> excel details", "extract-text -> xlsx"])
    else:
        unavailable_ops.append("inspect/extract-text -> xlsx (missing openpyxl)")

    if runtime.python_docx:
        available_ops.append("extract-text -> docx")
    else:
        unavailable_ops.append("extract-text -> docx (missing python-docx)")

    if runtime.pymupdf:
        available_ops.extend(["inspect -> pdf metadata", "extract-text -> pdf"])
    else:
        unavailable_ops.append("inspect/extract-text -> pdf (missing pymupdf)")

    return {
        "success": True,
        "python": sys.executable,
        "cwd": str(Path.cwd()),
        "runtime": {
            "libreoffice": runtime.libreoffice,
            "wkhtmltopdf": runtime.wkhtmltopdf,
            "openpyxl": runtime.openpyxl,
            "xlrd": runtime.xlrd,
            "markdown": runtime.markdown,
            "pdfkit": runtime.pdfkit,
            "python_docx": runtime.python_docx,
            "pymupdf": runtime.pymupdf,
        },
        "available_conversions": available_conversions,
        "unavailable_conversions": unavailable_conversions,
        "available_operations": available_ops,
        "unavailable_operations": unavailable_ops,
    }


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------


def normalize_extension(value: str) -> str:
    return value.lower().lstrip(".")


def _resolve_path(value: str) -> Path:
    return Path(value).expanduser().resolve()


def ensure_input_exists(input_path: str) -> Path:
    path = _resolve_path(input_path)
    if not path.is_file():
        raise FileOpsError(f"Input file not found: {path}")
    return path


def _ensure_path_exists(value: str) -> Path:
    path = _resolve_path(value)
    if not path.exists():
        raise FileOpsError(f"Input path not found: {path}")
    return path


def infer_output_path(input_path: Path, target_ext: str) -> Path:
    return input_path.with_suffix(f".{normalize_extension(target_ext)}")


def resolve_output_path(
    input_path: Path, output_path: str | None, target_ext: str
) -> Path:
    if output_path:
        output = _resolve_path(output_path)
    else:
        output = infer_output_path(input_path, target_ext).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    return output


def _format_timestamp(ts: float) -> str:
    return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Convert operations
# ---------------------------------------------------------------------------


def convert_image(input_path: Path, output_path: Path) -> None:
    from PIL import Image

    with Image.open(input_path) as image:
        target_ext = normalize_extension(output_path.suffix)
        target_format = "JPEG" if target_ext in {"jpg", "jpeg"} else target_ext.upper()
        save_image = image

        if target_ext in {"jpg", "jpeg"} and image.mode in {"RGBA", "LA", "P"}:
            background = Image.new("RGB", image.size, (255, 255, 255))
            alpha = image.getchannel("A") if "A" in image.getbands() else None
            background.paste(image.convert("RGBA"), mask=alpha)
            save_image = background

        save_image.save(output_path, format=target_format)


def convert_pdf_to_docx(input_path: Path, output_path: Path) -> None:
    from pdf2docx import Converter

    converter = Converter(str(input_path))
    try:
        converter.convert(str(output_path))
    finally:
        converter.close()


def _parse_sheet(value: str | None) -> str | int | None:
    if value is None:
        return 0
    if value.isdigit():
        return int(value)
    return value


def convert_excel_to_csv(
    input_path: Path, output_path: Path, sheet: str | None
) -> None:
    import pandas as pd

    try:
        dataframe = pd.read_excel(input_path, sheet_name=_parse_sheet(sheet))
    except ImportError as exc:
        raise FileOpsError(
            "Excel conversion requires openpyxl for .xlsx or xlrd for .xls."
        ) from exc

    if isinstance(dataframe, dict):
        raise FileOpsError(
            "Sheet selection returned multiple sheets; pass a single sheet name or index."
        )
    dataframe.to_csv(output_path, index=False)


def convert_docx_to_pdf(
    input_path: Path, output_path: Path, runtime: RuntimeStatus
) -> None:
    if not runtime.libreoffice:
        raise FileOpsError(
            "DOCX to PDF requires LibreOffice. Install soffice/libreoffice and retry."
        )

    with tempfile.TemporaryDirectory(prefix="docx-to-pdf-") as tmp:
        tmp_path = Path(tmp)
        source_copy = tmp_path / input_path.name
        shutil.copy2(input_path, source_copy)

        result = subprocess.run(
            [
                runtime.libreoffice,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(tmp_path),
                str(source_copy),
            ],
            capture_output=True,
            text=True,
            check=False,
        )
        generated = tmp_path / f"{source_copy.stem}.pdf"
        if result.returncode != 0 or not generated.exists():
            stderr = (
                result.stderr.strip()
                or result.stdout.strip()
                or "Unknown LibreOffice failure"
            )
            raise FileOpsError(f"LibreOffice conversion failed: {stderr}")

        shutil.move(str(generated), str(output_path))


def convert_html_like_to_pdf(
    input_path: Path, output_path: Path, runtime: RuntimeStatus
) -> None:
    if not runtime.wkhtmltopdf:
        raise FileOpsError("HTML/Markdown to PDF requires wkhtmltopdf.")
    if not runtime.pdfkit:
        raise FileOpsError("HTML/Markdown to PDF requires the pdfkit package.")

    import pdfkit

    source_path = input_path
    temp_html_path: Path | None = None

    if normalize_extension(input_path.suffix) in MARKDOWN_FORMATS:
        if not runtime.markdown:
            raise FileOpsError("Markdown to PDF requires the markdown package.")
        import markdown

        html_body = markdown.markdown(
            input_path.read_text(encoding="utf-8"),
            extensions=["extra", "tables", "fenced_code"],
        )
        html = (
            "<!doctype html><html><head><meta charset='utf-8'>"
            "<style>body{font-family:Arial,sans-serif;line-height:1.6;max-width:900px;"
            "margin:0 auto;padding:24px;}pre{background:#f5f5f5;padding:12px;overflow:auto;}"
            "code{background:#f5f5f5;padding:2px 4px;}table{border-collapse:collapse;width:100%;}"
            "th,td{border:1px solid #ddd;padding:8px;}</style></head><body>"
            f"{html_body}</body></html>"
        )
        tmp = tempfile.mkdtemp(prefix="md-to-pdf-")
        temp_html_path = Path(tmp) / f"{input_path.stem}.html"
        temp_html_path.write_text(html, encoding="utf-8")
        source_path = temp_html_path

    configuration = pdfkit.configuration(wkhtmltopdf=runtime.wkhtmltopdf)
    try:
        pdfkit.from_file(str(source_path), str(output_path), configuration=configuration)
    finally:
        if temp_html_path is not None:
            shutil.rmtree(temp_html_path.parent, ignore_errors=True)


def dispatch_conversion(
    input_path: Path, output_path: Path, runtime: RuntimeStatus, sheet: str | None
) -> None:
    source = normalize_extension(input_path.suffix)
    target = normalize_extension(output_path.suffix)

    if source in IMAGE_FORMATS and target in IMAGE_FORMATS:
        convert_image(input_path, output_path)
        return
    if source == "pdf" and target == "docx":
        convert_pdf_to_docx(input_path, output_path)
        return
    if source in EXCEL_FORMATS | LEGACY_EXCEL_FORMATS and target == "csv":
        convert_excel_to_csv(input_path, output_path, sheet)
        return
    if source == "docx" and target == "pdf":
        convert_docx_to_pdf(input_path, output_path, runtime)
        return
    if source in HTML_FORMATS | MARKDOWN_FORMATS and target == "pdf":
        convert_html_like_to_pdf(input_path, output_path, runtime)
        return

    raise FileOpsError(f"Unsupported conversion: {source} -> {target}")


# ---------------------------------------------------------------------------
# Inspect operations
# ---------------------------------------------------------------------------


def _inspect_image(input_path: Path) -> dict:
    from PIL import Image

    with Image.open(input_path) as image:
        return {
            "dimensions": {"width": image.width, "height": image.height},
            "color_mode": image.mode,
            "format": image.format,
        }


def _inspect_pdf(input_path: Path) -> dict:
    try:
        import fitz
    except ImportError as exc:
        raise FileOpsError("PDF inspection requires pymupdf (fitz).") from exc

    with fitz.open(str(input_path)) as doc:
        return {"page_count": doc.page_count}


def _inspect_excel(input_path: Path) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise FileOpsError("Excel inspection requires openpyxl.") from exc

    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheets = list(wb.sheetnames)
        row_counts = {}
        for name in sheets:
            ws = wb[name]
            row_counts[name] = sum(
                1
                for row in ws.iter_rows(values_only=True)
                if any(cell is not None for cell in row)
            )
        return {"sheet_names": sheets, "row_counts": row_counts}
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Archive operations
# ---------------------------------------------------------------------------


def _add_to_archive(
    archive: zipfile.ZipFile, path: Path, prefix: str = ""
) -> int:
    """Add a file or directory tree to *archive*. Return files added.

    *prefix* is prepended to avoid basename collisions when archiving
    multiple inputs that share the same filename.
    """
    if path.is_file():
        arcname = f"{prefix}{path.name}" if prefix else path.name
        archive.write(path, arcname=arcname)
        return 1

    count = 0
    base = f"{prefix}{path.name}" if prefix else path.name
    for child in sorted(path.rglob("*")):
        if child.is_file():
            arcname = str(Path(base) / child.relative_to(path))
            archive.write(child, arcname=arcname)
            count += 1
    return count


def _safe_extract(archive: zipfile.ZipFile, dest: Path) -> int:
    """Extract archive with ZipSlip protection. Return entries extracted."""
    dest = dest.resolve()
    for member in archive.namelist():
        target = (dest / member).resolve()
        if not target.is_relative_to(dest):
            raise FileOpsError(
                f"Blocked path traversal in archive entry: {member}"
            )
    archive.extractall(dest)
    return len(archive.namelist())


# ---------------------------------------------------------------------------
# Text extraction operations
# ---------------------------------------------------------------------------


def _extract_text_pdf(input_path: Path) -> str:
    try:
        import fitz
    except ImportError as exc:
        raise FileOpsError("PDF text extraction requires pymupdf (fitz).") from exc

    with fitz.open(str(input_path)) as doc:
        return "\n".join(page.get_text().strip() for page in doc).strip()


def _extract_text_docx(input_path: Path) -> str:
    try:
        from docx import Document
    except ImportError as exc:
        raise FileOpsError("DOCX text extraction requires python-docx.") from exc

    document = Document(str(input_path))
    return "\n".join(p.text for p in document.paragraphs if p.text).strip()


def _extract_text_xlsx(input_path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise FileOpsError("XLSX text extraction requires openpyxl.") from exc

    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        chunks: list[str] = []
        for ws in wb.worksheets:
            lines = [f"[Sheet: {ws.title}]"]
            for row in ws.iter_rows(values_only=True):
                values = [str(c) for c in row if c is not None]
                if values:
                    lines.append("\t".join(values))
            chunks.append("\n".join(lines))
        return "\n\n".join(chunks).strip()
    finally:
        wb.close()


def _extract_text_html(input_path: Path) -> str:
    extractor = _HTMLTextExtractor()
    extractor.feed(input_path.read_text(encoding="utf-8"))
    extractor.close()
    return extractor.get_text()


def _extract_text_markdown(input_path: Path) -> str:
    return input_path.read_text(encoding="utf-8")


def _extract_text_image(input_path: Path) -> str:
    from PIL import ExifTags, Image

    with Image.open(input_path) as image:
        exif = image.getexif()
        if not exif:
            return "No EXIF metadata found."
        lines = []
        for key, value in sorted(exif.items()):
            tag = ExifTags.TAGS.get(key, str(key))
            lines.append(f"{tag}: {value}")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Command handlers
# ---------------------------------------------------------------------------


def handle_convert(args: argparse.Namespace) -> dict:
    runtime = detect_runtime()
    input_path = ensure_input_exists(args.input)
    output_ext = normalize_extension(args.to)
    output_path = resolve_output_path(input_path, args.output, output_ext)

    if output_path == input_path:
        raise FileOpsError("Output path must differ from input path.")
    if output_path.exists() and not args.overwrite:
        raise FileOpsError(
            f"Output already exists: {output_path}. Pass --overwrite to replace."
        )

    dispatch_conversion(input_path, output_path, runtime, args.sheet)

    return {
        "success": True,
        "input_path": str(input_path),
        "output_path": str(output_path),
        "conversion": f"{normalize_extension(input_path.suffix)} -> {output_ext}",
    }


def handle_inspect(args: argparse.Namespace) -> dict:
    input_path = ensure_input_exists(args.input)
    ext = normalize_extension(input_path.suffix)
    st = input_path.stat()

    result: dict = {
        "success": True,
        "input_path": str(input_path),
        "size_bytes": st.st_size,
        "mime_type": mimetypes.guess_type(str(input_path))[0] or "application/octet-stream",
        "last_modified": _format_timestamp(st.st_mtime),
        "permissions": stat.filemode(st.st_mode),
    }

    if ext in IMAGE_FORMATS:
        result.update(_inspect_image(input_path))
    elif ext == "pdf":
        result.update(_inspect_pdf(input_path))
    elif ext in EXCEL_FORMATS:
        result.update(_inspect_excel(input_path))

    return result


def handle_archive(args: argparse.Namespace) -> dict:
    input_paths = [_ensure_path_exists(v) for v in args.input]
    output_path = _resolve_path(args.output)

    if args.extract:
        if len(input_paths) != 1:
            raise FileOpsError("Extraction requires exactly one .zip input.")
        src = input_paths[0]
        if not src.is_file() or not zipfile.is_zipfile(src):
            raise FileOpsError(f"Not a valid zip file: {src}")
        if output_path.exists() and not output_path.is_dir():
            raise FileOpsError(f"Extraction target must be a directory: {output_path}")

        output_path.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(src) as archive:
            entries = _safe_extract(archive, output_path)

        return {
            "success": True,
            "action": "extract",
            "input_path": str(src),
            "output_path": str(output_path),
            "entries": entries,
        }

    # Create archive
    if output_path.suffix.lower() != ".zip":
        raise FileOpsError("Archive output must use .zip extension.")
    if output_path.exists():
        raise FileOpsError(f"Output already exists: {output_path}")
    for p in input_paths:
        if p == output_path:
            raise FileOpsError("Output path must differ from input paths.")
        if p.is_dir() and output_path.is_relative_to(p):
            raise FileOpsError(
                f"Archive output cannot be inside input directory: {p}"
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    total = 0
    # Detect basename collisions and prefix with parent dir name
    names = [p.name for p in input_paths]
    has_collision = len(names) != len(set(names))
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for p in input_paths:
            prefix = f"{p.parent.name}/" if has_collision and p.is_file() else ""
            total += _add_to_archive(archive, p, prefix=prefix)

    return {
        "success": True,
        "action": "create",
        "input_paths": [str(p) for p in input_paths],
        "output_path": str(output_path),
        "files_added": total,
    }


def handle_extract_text(args: argparse.Namespace) -> dict:
    input_path = ensure_input_exists(args.input)
    ext = normalize_extension(input_path.suffix)

    extractors: dict[str, callable] = {
        "pdf": _extract_text_pdf,
        "docx": _extract_text_docx,
        "html": _extract_text_html,
        "htm": _extract_text_html,
        "md": _extract_text_markdown,
        "markdown": _extract_text_markdown,
    }
    for fmt in IMAGE_FORMATS:
        extractors[fmt] = _extract_text_image
    for fmt in EXCEL_FORMATS:
        extractors[fmt] = _extract_text_xlsx

    extractor = extractors.get(ext)
    if extractor is None:
        raise FileOpsError(f"Unsupported text extraction format: {ext}")

    return {
        "success": True,
        "text": extractor(input_path),
        "source": str(input_path),
        "format": ext,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Local file operations skill entry point."
    )
    sub = parser.add_subparsers(dest="command", required=True)

    sub.add_parser("health", help="Report available operations and dependencies.")

    convert_p = sub.add_parser("convert", help="Convert a local file.")
    convert_p.add_argument("--input", required=True, help="Input file path.")
    convert_p.add_argument("--to", required=True, help="Target extension (jpg, docx, csv, pdf).")
    convert_p.add_argument("--output", help="Explicit output file path.")
    convert_p.add_argument("--sheet", help="Excel sheet name or zero-based index.")
    convert_p.add_argument("--overwrite", action="store_true", help="Replace existing output.")

    inspect_p = sub.add_parser("inspect", help="Inspect a file and return metadata.")
    inspect_p.add_argument("--input", required=True, help="Input file path.")

    archive_p = sub.add_parser("archive", help="Create or extract a zip archive.")
    archive_p.add_argument("--input", required=True, nargs="+", help="Input paths.")
    archive_p.add_argument("--output", required=True, help="Zip path or extraction directory.")
    archive_p.add_argument("--extract", action="store_true", help="Extract zip to output directory.")

    extract_p = sub.add_parser("extract-text", help="Extract text content from a file.")
    extract_p.add_argument("--input", required=True, help="Input file path.")

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    try:
        handlers = {
            "health": lambda: build_health_report(),
            "convert": lambda: handle_convert(args),
            "inspect": lambda: handle_inspect(args),
            "archive": lambda: handle_archive(args),
            "extract-text": lambda: handle_extract_text(args),
        }
        handler = handlers.get(args.command)
        if handler is None:
            raise FileOpsError(f"Unknown command: {args.command}")
        result = handler()
    except FileOpsError as exc:
        result = {"success": False, "error": str(exc)}
        print(json.dumps(result, indent=2, sort_keys=True))
        return 1
    except Exception as exc:
        result = {"success": False, "error": f"Unexpected error: {exc}"}
        print(json.dumps(result, indent=2, sort_keys=True))
        return 1

    print(json.dumps(result, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
